import openpyxl


def get_rows_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    rows = sheet.max_row
    return (rows)


def get_column_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    column = sheet.max_column
    return column


def read_data(file, sheet_name, row_number, column_number):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.cell(row=row_number, column=column_number).value


def write_data(file, sheet_name, row_number, column_number, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    sheet.cell(row=row_number, column=column_number).value = data
    workbook.save(file)

# path = "C:\\Users\\HP\\Documents\\Book1.xlsx"
